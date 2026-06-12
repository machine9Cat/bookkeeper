"""报表导出模块 - Excel 和 PDF"""
import os
from datetime import datetime
from fastapi.responses import FileResponse


async def generate_monthly_report(data: dict, format: str) -> FileResponse:
    """生成月度报表"""
    if format == "excel":
        return await generate_excel_report(data)
    elif format == "pdf":
        return await generate_pdf_report(data)
    else:
        raise ValueError(f"不支持的格式: {format}")


async def generate_excel_report(data: dict) -> FileResponse:
    """生成 Excel 报表"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("请安装 openpyxl: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = f"{data['month']} 月度采购报表"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, size=16)
    money_format = '#,##0.00'
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    expenses = data.get('expenses', [])
    total = data.get('total_expense', 0)

    # 标题
    ws.merge_cells('A1:R1')
    ws['A1'] = f"📦 {data['month']} 月度采购支出报表"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = f"用户: {data['username']}"
    ws['A3'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    # 汇总
    ws['A5'] = "支出汇总"
    ws['A5'].font = Font(bold=True, size=14)
    ws['A6'] = "总支出"
    ws['B6'] = total
    ws['B6'].number_format = money_format
    ws['A7'] = "记录条数"
    ws['B7'] = len(expenses)

    # 明细表头
    row = 9
    headers = [
        "日期", "大类", "物品名称", "规格型号", "数量", "单位",
        "单价", "总价", "经手人", "付款方式", "购买渠道",
        "入库", "入账", "发票", "发票类型", "记账月日", "备注"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border
    row += 1

    for exp in expenses:
        ws.cell(row=row, column=1, value=exp.get('expense_date', ''))
        ws.cell(row=row, column=2, value=exp.get('category_name', ''))
        ws.cell(row=row, column=3, value=exp.get('item_name', ''))
        ws.cell(row=row, column=4, value=exp.get('specification', ''))
        ws.cell(row=row, column=5, value=exp.get('quantity', 0))
        ws.cell(row=row, column=6, value=exp.get('unit', ''))
        ws.cell(row=row, column=7, value=exp.get('unit_price', 0))
        ws.cell(row=row, column=7).number_format = money_format
        ws.cell(row=row, column=8, value=exp.get('total_price', 0))
        ws.cell(row=row, column=8).number_format = money_format
        ws.cell(row=row, column=9, value=exp.get('handler', ''))
        ws.cell(row=row, column=10, value=exp.get('payment_method', ''))
        ws.cell(row=row, column=11, value=exp.get('purchase_channel', ''))
        ws.cell(row=row, column=12, value=exp.get('inventory_in', ''))
        ws.cell(row=row, column=13, value=exp.get('booked', ''))
        ws.cell(row=row, column=14, value=exp.get('invoice', ''))
        ws.cell(row=row, column=15, value=exp.get('invoice_type', ''))
        ws.cell(row=row, column=16, value=exp.get('accounting_date', ''))
        ws.cell(row=row, column=17, value=exp.get('note', ''))

        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = border
        row += 1

    # 列宽
    col_widths = [12, 12, 18, 14, 8, 6, 10, 12, 10, 12, 12, 6, 6, 6, 10, 12, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    output_dir = os.path.join(os.path.dirname(__file__), "..", "data", "exports")
    os.makedirs(output_dir, exist_ok=True)
    filename = f"report_{data['month']}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)

    return FileResponse(
        filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


async def generate_pdf_report(data: dict) -> FileResponse:
    """生成 PDF 报表"""
    try:
        from weasyprint import HTML
    except ImportError:
        raise ImportError("请安装 weasyprint: pip install weasyprint")

    expenses = data.get('expenses', [])
    total = data.get('total_expense', 0)
    rows_html = ""
    for exp in expenses:
        rows_html += f"""
                <tr>
                    <td>{exp.get('expense_date', '')}</td>
                    <td>{exp.get('category_name', '')}</td>
                    <td>{exp.get('item_name', '')}</td>
                    <td>{exp.get('specification', '')}</td>
                    <td>{exp.get('quantity', '')}</td>
                    <td>{exp.get('unit', '')}</td>
                    <td>¥{exp.get('unit_price', 0):,.2f}</td>
                    <td>¥{exp.get('total_price', 0):,.2f}</td>
                    <td>{exp.get('handler', '')}</td>
                </tr>
        """

    html_content = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><style>
    body {{ font-family: 'Microsoft YaHei', sans-serif; margin: 30px; font-size: 12px; }}
    h1 {{ color: #2c3e50; text-align: center; font-size: 18px; }}
    .summary {{ background: #f0f4f8; padding: 15px; border-radius: 6px; margin: 15px 0; text-align: center; }}
    .amount {{ font-size: 22px; font-weight: bold; color: #e74c3c; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
    th {{ background: #4472C4; color: white; padding: 8px 6px; text-align: left; font-size: 11px; }}
    td {{ padding: 6px; border-bottom: 1px solid #ddd; }}
    tr:nth-child(even) {{ background: #f8f9fa; }}
    .footer {{ text-align: center; margin-top: 30px; color: #999; font-size: 11px; }}
</style></head><body>
    <h1>📦 {data['month']} 月度采购支出报表</h1>
    <p style="text-align:center;">用户: {data['username']} | 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
    <div class="summary">
        <div>月度总支出</div>
        <div class="amount">¥{total:,.2f}</div>
        <div style="margin-top:4px;color:#666;">共 {len(expenses)} 条记录</div>
    </div>
    <table>
        <thead><tr>
            <th>日期</th><th>大类</th><th>物品名称</th><th>规格型号</th>
            <th>数量</th><th>单位</th><th>单价</th><th>总价</th><th>经手人</th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
    </table>
    <div class="footer"><p>此报表由采购支出记账系统自动生成</p></div>
</body></html>"""

    output_dir = os.path.join(os.path.dirname(__file__), "..", "data", "exports")
    os.makedirs(output_dir, exist_ok=True)
    filename = f"report_{data['month']}.pdf"
    filepath = os.path.join(output_dir, filename)

    HTML(string=html_content).write_pdf(filepath)

    return FileResponse(
        filepath,
        filename=filename,
        media_type="application/pdf"
    )
